from fastapi import FastAPI, APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import json
import re
import uuid
import logging
import requests
from pathlib import Path
import bcrypt
from pydantic import BaseModel, Field, ConfigDict, BeforeValidator
from typing import List, Optional, Annotated
from datetime import datetime, timezone, timedelta

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

STATIC_DIR = ROOT_DIR / 'static'
MODEL_PATH = STATIC_DIR / 'nab3d.glb'
OBJECTS_CACHE = STATIC_DIR / 'objects.json'
VIEWER_PATH = STATIC_DIR / 'viewer.html'

VALID_STATUSES = {"fabricado", "enviado", "instalado", "entregable", "observaciones"}
STATUS_ORDER = ["fabricado", "enviado", "instalado", "entregable", "observaciones"]
STATUS_LABELS = {
    "fabricado": "Fabricado",
    "enviado": "Enviado",
    "instalado": "Instalado",
    "entregado": "Entregado",  # kept for historical events in old reports
    "entregable": "Entregable",
    "observaciones": "Observaciones",
    "all": "Todas",
}

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ---------- Object storage (Emergent) ----------

STORAGE_BASE = (os.environ.get("INTEGRATION_PROXY_URL") or "").strip() or "https://integrations.emergentagent.com"
STORAGE_URL = STORAGE_BASE.rstrip("/") + "/objstore/api/v1/storage"
APP_NAME = "bimtracker"
storage_key = None


def init_storage(force: bool = False):
    global storage_key
    if storage_key and not force:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": os.environ.get("EMERGENT_LLM_KEY")}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key


def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                        headers={"X-Storage-Key": key, "Content-Type": content_type},
                        data=data, timeout=120)
    if resp.status_code == 404:
        key = init_storage(force=True)
        resp = requests.put(f"{STORAGE_URL}/objects/{path}",
                            headers={"X-Storage-Key": key, "Content-Type": content_type},
                            data=data, timeout=120)
    resp.raise_for_status()
    return resp.json()


def storage_get_object(path: str):
    key = init_storage()
    resp = requests.get(f"{STORAGE_URL}/objects/{path}", headers={"X-Storage-Key": key}, timeout=60)
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

# ---------- Object catalog (extracted from GLB) ----------

def load_objects():
    if OBJECTS_CACHE.exists():
        with open(OBJECTS_CACHE) as f:
            return json.load(f)
    from pygltflib import GLTF2
    g = GLTF2().load(str(MODEL_PATH))
    objs, seen = [], set()
    for n in g.nodes:
        if n.mesh is not None and n.name and n.name not in seen:
            seen.add(n.name)
            objs.append({'name': n.name, 'mark': n.name.split(' ')[0]})
    with open(OBJECTS_CACHE, 'w') as f:
        json.dump(objs, f)
    return objs


OBJECTS: List[dict] = []
NAME_SET = set()
FACADE_NAMES = set()
FACADES: dict = {}
FACADES_PATH = STATIC_DIR / 'facades.json'
DIMS: dict = {}
DIMS_PATH = STATIC_DIR / 'dims.json'
VALID_FACADES = {"norte", "sur", "este", "oeste"}
FACADE_LABELS = {"norte": "Norte", "sur": "Sur", "este": "Este", "oeste": "Oeste", "all": "Todas"}

# Façade = marks starting with C or L followed by a number (C1, C7B-2, L3G, ...) → 533 objects (per user)
FACADE_RE = re.compile(r"^[CL]\d")


def is_facade(mark: str) -> bool:
    return bool(FACADE_RE.match(mark))


@app.on_event("startup")
async def startup():
    global OBJECTS, NAME_SET, FACADE_NAMES, FACADES, DIMS
    OBJECTS = load_objects()
    NAME_SET = {o['name'] for o in OBJECTS}
    FACADE_NAMES = {o['name'] for o in OBJECTS if is_facade(o['mark'])}
    if FACADES_PATH.exists():
        with open(FACADES_PATH) as f:
            FACADES = json.load(f)
    if DIMS_PATH.exists():
        with open(DIMS_PATH) as f:
            DIMS = json.load(f)
    await db.tags.create_index("object_name", unique=True)
    # 'entregado' status removed from the app: clear it from existing tags (history is kept)
    await db.tags.update_many({"status": "entregado"}, {"$set": {"status": None}})
    # Backfill created_at / history for tags created before history support
    async for doc in db.tags.find():
        upd = {}
        date = doc.get("updated_at") or datetime.now(timezone.utc).isoformat()
        if not doc.get("created_at"):
            upd["created_at"] = date
        if doc.get("status") and not doc.get("history"):
            upd["history"] = [{"status": doc["status"], "date": date}]
        if upd:
            await db.tags.update_one({"_id": doc["_id"]}, {"$set": upd})
    try:
        init_storage()
        logging.info("Object storage initialized")
    except Exception as e:
        logging.error(f"Storage init failed: {e}")
    logging.info(f"Loaded {len(OBJECTS)} objects ({len(FACADE_NAMES)} facade) from model")


# ---------- Models ----------

PyObjectId = Annotated[str, BeforeValidator(str)]


class BaseDocument(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    id: Optional[PyObjectId] = Field(default=None, alias="_id")

    def to_mongo(self) -> dict:
        d = self.model_dump(by_alias=True)
        d.pop("_id", None)
        return d

    @classmethod
    def from_mongo(cls, doc: Optional[dict]):
        return cls(**doc) if doc else None


class Tag(BaseDocument):
    object_name: str
    status: Optional[str] = None
    observation: str = ""
    updated_at: str = ""
    created_at: str = ""
    history: List[dict] = []
    observations: List[dict] = []


class TagUpsert(BaseModel):
    object_name: str
    status: Optional[str] = None
    observation: str = ""
    photo: Optional[str] = None


class AdminVerifyRequest(BaseModel):
    password: str


class FacadesPayload(BaseModel):
    facades: dict


class DimsPayload(BaseModel):
    dims: dict


# ---------- Helpers ----------

def display_name(name: str) -> str:
    parts = name.split(" ")
    if len(parts) >= 2 and parts[0] == parts[1]:
        parts.pop(1)
    return " ".join(parts)


async def fetch_tags_map() -> dict:
    tags = {}
    async for doc in db.tags.find():
        try:
            t = Tag.from_mongo(doc)
        except Exception:
            logging.warning(f"Skipping malformed tag doc: {doc.get('object_name')}")
            continue
        tags[t.object_name] = t
    return tags


# ---------- Routes ----------

@api_router.get("/")
async def root():
    return {"message": "BIMTracker API", "objects": len(OBJECTS)}


@api_router.get("/model")
async def get_model():
    return FileResponse(MODEL_PATH, media_type="model/gltf-binary",
                        headers={"Cache-Control": "public, max-age=86400"})


@api_router.get("/viewer")
async def get_viewer():
    return FileResponse(VIEWER_PATH, media_type="text/html",
                        headers={"Cache-Control": "no-cache"})


@api_router.post("/admin/verify")
async def verify_admin(payload: AdminVerifyRequest):
    """Verify the shared admin password against the bcrypt hash in the env."""
    hashed = os.environ.get("ADMIN_PASSWORD_HASH", "")
    ok = False
    if hashed and payload.password:
        try:
            ok = bcrypt.checkpw(payload.password.encode(), hashed.encode())
        except Exception:
            ok = False
    if not ok:
        return {"ok": False, "message": "Contraseña de administrador incorrecta."}
    return {"ok": True, "message": "Acceso de administrador concedido."}


@api_router.get("/facades/count")
async def facades_count():
    return {"count": len(FACADES)}


@api_router.post("/facades")
async def save_facades(payload: FacadesPayload):
    """Persist the per-object cardinal orientation computed by the 3D viewer."""
    global FACADES
    clean = {k: v for k, v in payload.facades.items() if k in NAME_SET and v in VALID_FACADES}
    if not clean:
        raise HTTPException(status_code=422, detail="Sin orientaciones válidas")
    FACADES = clean
    with open(FACADES_PATH, 'w') as f:
        json.dump(clean, f)
    return {"saved": len(clean)}


@api_router.get("/dims/count")
async def dims_count():
    return {"count": len(DIMS)}


@api_router.post("/dims")
async def save_dims(payload: DimsPayload):
    """Persist per-object bounding box sizes [sx, sy, sz] computed by the 3D viewer."""
    global DIMS
    clean = {}
    for k, v in payload.dims.items():
        if k in NAME_SET and isinstance(v, list) and len(v) == 3:
            try:
                clean[k] = [round(float(x), 4) for x in v]
            except (TypeError, ValueError):
                continue
    if not clean:
        raise HTTPException(status_code=422, detail="Sin dimensiones válidas")
    DIMS = clean
    with open(DIMS_PATH, 'w') as f:
        json.dump(clean, f)
    return {"saved": len(clean)}


ALLOWED_IMG = {"image/jpeg", "image/png", "image/webp", "image/gif", "image/heic", "image/heif"}


@api_router.post("/upload")
async def upload_photo(file: UploadFile = File(...)):
    if file.content_type not in ALLOWED_IMG:
        raise HTTPException(status_code=422, detail="Solo se permiten imágenes (JPG, PNG, WEBP, GIF)")
    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="La imagen supera el límite de 10 MB")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else "jpg"
    path = f"{APP_NAME}/uploads/{uuid.uuid4()}.{ext}"
    try:
        result = put_object(path, data, file.content_type)
    except Exception as e:
        logging.error(f"Photo upload failed: {e}")
        raise HTTPException(status_code=502, detail="No se pudo subir la foto. Inténtalo de nuevo.")
    await db.files.insert_one({
        "storage_path": result["path"],
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": result.get("size", len(data)),
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"path": result["path"]}


@api_router.get("/files/{path:path}")
async def serve_file(path: str):
    record = await db.files.find_one({"storage_path": path, "is_deleted": False})
    if not record:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    try:
        data, ct = storage_get_object(path)
    except Exception as e:
        logging.error(f"Photo download failed: {e}")
        raise HTTPException(status_code=502, detail="No se pudo descargar la foto.")
    return Response(content=data, media_type=record.get("content_type") or ct,
                    headers={"Cache-Control": "public, max-age=86400"})


@api_router.get("/photos")
async def list_photos(
    facade: str = "all",
    from_: str = Query(default="", alias="from"),
    to: str = "",
):
    """Gallery of all obra photos attached to observations, filterable by facade and date."""
    if facade != "all" and facade not in VALID_FACADES:
        raise HTTPException(status_code=422, detail="Fachada inválida")
    items = []
    async for doc in db.tags.find({"observations.photo": {"$exists": True}}):
        t = Tag.from_mongo(doc)
        fac = FACADES.get(t.object_name)
        if facade != "all" and fac != facade:
            continue
        for ob in t.observations or []:
            photo = ob.get("photo")
            if not photo:
                continue
            d = (ob.get("date") or "")[:10]
            if from_ and d < from_:
                continue
            if to and d > to:
                continue
            items.append({
                "name": t.object_name,
                "mark": t.object_name.split(" ")[0],
                "facade": fac,
                "photo": photo,
                "text": ob.get("text", ""),
                "date": ob.get("date"),
                "status": t.status,
            })
    items.sort(key=lambda x: x["date"] or "", reverse=True)
    return {"total": len(items), "items": items}


@api_router.get("/objects")
async def list_objects(
    search: str = "",
    status: str = "all",
    facade: str = "all",
    skip: int = 0,
    limit: int = Query(default=50, le=200),
):
    tags = await fetch_tags_map()
    q = search.strip().lower()
    filtered = []
    for o in OBJECTS:
        if q and q not in o['name'].lower():
            continue
        if facade != "all" and FACADES.get(o['name']) != facade:
            continue
        t = tags.get(o['name'])
        t_status = t.status if t else None
        if status == "none":
            if t_status is not None:
                continue
        elif status != "all":
            if t_status != status:
                continue
        filtered.append({
            "name": o['name'],
            "mark": o['mark'],
            "facade": FACADES.get(o['name']),
            "status": t_status,
            "observation": t.observation if t else "",
        })
    total = len(filtered)
    return {"total": total, "items": filtered[skip:skip + limit]}


@api_router.get("/object")
async def get_object(name: str):
    if name not in NAME_SET:
        raise HTTPException(status_code=404, detail="Objeto no encontrado")
    doc = await db.tags.find_one({"object_name": name})
    t = Tag.from_mongo(doc)
    mark = name.split(' ')[0]
    return {
        "name": name,
        "mark": mark,
        "facade": FACADES.get(name),
        "dimensions": DIMS.get(name),
        "status": t.status if t else None,
        "observation": t.observation if t else "",
        "created_at": t.created_at if t else "",
        "history": t.history if t else [],
        "observations": t.observations if t else [],
    }


@api_router.get("/tags")
async def get_tags():
    tags = await fetch_tags_map()
    return {
        name: {"status": t.status, "observation": t.observation}
        for name, t in tags.items()
    }


@api_router.put("/tags")
async def upsert_tag(payload: TagUpsert):
    if payload.object_name not in NAME_SET:
        raise HTTPException(status_code=404, detail="Objeto no encontrado")
    if payload.status is not None and payload.status not in VALID_STATUSES:
        raise HTTPException(status_code=422, detail="Estado inválido")
    obs = payload.observation.strip()
    photo = (payload.photo or "").strip() or None
    existing = await db.tags.find_one({"object_name": payload.object_name})
    prev = Tag.from_mongo(existing) if existing else None
    # delete only if clearing status, no new observation/photo and no observation history
    if payload.status is None and not obs and not photo and not (prev and prev.observations):
        await db.tags.delete_one({"object_name": payload.object_name})
        return {"object_name": payload.object_name, "status": None, "observation": "", "observations": []}
    now = datetime.now(timezone.utc).isoformat()
    created_at = prev.created_at if prev and prev.created_at else now
    history = list(prev.history) if prev else []
    if payload.status and (not prev or prev.status != payload.status):
        history.append({"status": payload.status, "date": now})
    observations = list(prev.observations) if prev else []
    if (obs or photo) and (photo or not observations or observations[-1].get("text") != obs):
        entry = {"text": obs, "date": now}
        if photo:
            entry["photo"] = photo
        observations.append(entry)
    latest_obs = observations[-1]["text"] if observations else ""
    tag = Tag(
        object_name=payload.object_name,
        status=payload.status,
        observation=latest_obs,
        updated_at=now,
        created_at=created_at,
        history=history,
        observations=observations,
    )
    await db.tags.update_one(
        {"object_name": payload.object_name},
        {"$set": tag.to_mongo()},
        upsert=True,
    )
    return {
        "object_name": tag.object_name,
        "status": tag.status,
        "observation": tag.observation,
        "created_at": tag.created_at,
        "history": tag.history,
        "observations": tag.observations,
    }


@api_router.get("/stats")
async def get_stats():
    tags = await fetch_tags_map()
    counts = {s: 0 for s in STATUS_ORDER}
    con_obs = 0
    tagged = 0
    for name, t in tags.items():
        if name not in FACADE_NAMES:
            continue
        if t.status in counts:
            counts[t.status] += 1
            tagged += 1
        if t.observation:
            con_obs += 1
    total = len(FACADE_NAMES)
    por_fachada = {d: {"total": 0, "etiquetados": 0} for d in ["norte", "sur", "este", "oeste"]}
    for name in FACADE_NAMES:
        d = FACADES.get(name)
        if d not in por_fachada:
            continue
        por_fachada[d]["total"] += 1
        t = tags.get(name)
        if t and t.status in counts:
            por_fachada[d]["etiquetados"] += 1
    # weekly installed comparison (Mon-Sun, facade panels, from history events)
    today = datetime.now(timezone.utc).date()
    monday = today - timedelta(days=today.weekday())
    cur_from, cur_to = monday.isoformat(), (monday + timedelta(days=6)).isoformat()
    prev_from, prev_to = (monday - timedelta(days=7)).isoformat(), (monday - timedelta(days=1)).isoformat()
    semana = {"actual": 0, "anterior": 0, "desde": cur_from, "hasta": cur_to}
    for name, t in tags.items():
        if name not in FACADE_NAMES or t.status != "instalado":
            continue
        inst_dates = [(ev.get("date") or "")[:10] for ev in t.history or [] if ev.get("status") == "instalado"]
        if not inst_dates:
            continue
        d = max(inst_dates)
        if cur_from <= d <= cur_to:
            semana["actual"] += 1
        elif prev_from <= d <= prev_to:
            semana["anterior"] += 1
    return {
        "total": total,
        "counts": counts,
        "etiquetados": tagged,
        "sin_estado": total - tagged,
        "con_observaciones": con_obs,
        "por_fachada": por_fachada,
        "semana": semana,
    }


async def build_report(from_: str, to: str, status: str, facade: str = "all"):
    if status != "all" and status not in VALID_STATUSES:
        raise HTTPException(status_code=422, detail="Estado inválido")
    if facade != "all" and facade not in VALID_FACADES:
        raise HTTPException(status_code=422, detail="Fachada inválida")
    tags = await fetch_tags_map()
    counts = {s: 0 for s in STATUS_ORDER}
    items = []
    for name, t in tags.items():
        if facade != "all" and FACADES.get(name) != facade:
            continue
        for ev in t.history or []:
            ev_status = ev.get("status")
            d = (ev.get("date") or "")[:10]
            if not ev_status or not d:
                continue
            if from_ and d < from_:
                continue
            if to and d > to:
                continue
            if status != "all" and ev_status != status:
                continue
            items.append({
                "name": name,
                "mark": name.split(' ')[0],
                "facade": FACADES.get(name),
                "status": ev_status,
                "date": ev.get("date"),
                "observation": t.observation,
            })
            if ev_status in counts:
                counts[ev_status] += 1
    items.sort(key=lambda x: x["date"] or "", reverse=True)
    return {"total": len(items), "counts": counts, "items": items}


@api_router.get("/report")
async def get_report(
    from_: str = Query(default="", alias="from"),
    to: str = "",
    status: str = "all",
    facade: str = "all",
):
    """Report of status-change events (from tag history) within a date range.
    from/to: YYYY-MM-DD (inclusive). status: 'all' or a status key. facade: 'all'|norte|sur|este|oeste."""
    return await build_report(from_, to, status, facade)


def make_pdf(data: dict, from_: str, to: str, status: str, facade: str = "all") -> bytes:
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="Reporte BIMTracker",
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elems = []
    # Corporate letterhead with the 3 logos
    logo_defs = [
        ("logo_fiberkret.png", 1600 / 533),
        ("logo_entrepisos.png", 921 / 371),
        ("logo_grcontreras.png", 921 / 372),
    ]
    lh = 10 * mm
    imgs = [RLImage(str(STATIC_DIR / f), width=lh * r, height=lh)
            for f, r in logo_defs if (STATIC_DIR / f).exists()]
    if imgs:
        letterhead = Table([imgs], colWidths=[i.drawWidth + 8 * mm for i in imgs], hAlign="LEFT")
        letterhead.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elems.append(letterhead)
        elems.append(Spacer(1, 5 * mm))
    period = f"Período: {from_} a {to} &nbsp;·&nbsp; Etiqueta: {STATUS_LABELS.get(status, status)}"
    if facade != "all":
        period += f" &nbsp;·&nbsp; Fachada: {FACADE_LABELS.get(facade, facade)}"
    elems += [
        Paragraph("Reporte BIMTracker — Fachada", styles["Title"]),
        Paragraph(period, styles["Normal"]),
        Spacer(1, 4 * mm),
    ]
    summary = "   ·   ".join(
        f"{STATUS_LABELS[k]}: {v}" for k, v in data["counts"].items() if v
    )
    elems.append(Paragraph(f"Total de movimientos: {data['total']}", styles["Heading3"]))
    if summary:
        elems.append(Paragraph(summary, styles["Normal"]))
    elems.append(Spacer(1, 5 * mm))
    rows = [["Pieza", "Estado", "Fachada", "Fecha", "Observación"]]
    for it in data["items"]:
        rows.append([
            Paragraph(display_name(it["name"]), styles["BodyText"]),
            STATUS_LABELS.get(it["status"], it["status"]),
            FACADE_LABELS.get(it.get("facade") or "", "—"),
            (it["date"] or "")[:10],
            Paragraph(it["observation"] or "", styles["BodyText"]),
        ])
    table = Table(rows, colWidths=[52 * mm, 22 * mm, 20 * mm, 22 * mm, 54 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1C1C1E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#C7C7CC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F2F2F7")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elems.append(table)
    doc.build(elems)
    return buf.getvalue()


def make_xlsx(data: dict, from_: str, to: str, status: str, facade: str = "all") -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.append(["Reporte BIMTracker — Fachada"])
    ws["A1"].font = Font(bold=True, size=14)
    row2 = [f"Período: {from_} a {to}", f"Etiqueta: {STATUS_LABELS.get(status, status)}"]
    if facade != "all":
        row2.append(f"Fachada: {FACADE_LABELS.get(facade, facade)}")
    ws.append(row2)
    ws.append([f"Total de movimientos: {data['total']}"])
    ws.append([f"{STATUS_LABELS[k]}: {v}" for k, v in data["counts"].items() if v])
    ws.append([])
    header = ["Pieza", "Estado", "Fachada", "Fecha", "Observación"]
    ws.append(header)
    hrow = ws.max_row
    for col in range(1, len(header) + 1):
        c = ws.cell(row=hrow, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1C1C1E", end_color="1C1C1E", fill_type="solid")
    for it in data["items"]:
        ws.append([
            display_name(it["name"]),
            STATUS_LABELS.get(it["status"], it["status"]),
            FACADE_LABELS.get(it.get("facade") or "", "—"),
            (it["date"] or "")[:10],
            it["observation"] or "",
        ])
    for col, width in zip("ABCDE", [40, 16, 12, 14, 50]):
        ws.column_dimensions[col].width = width
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@api_router.get("/report/export")
async def export_report(
    format: str = "xlsx",
    from_: str = Query(default="", alias="from"),
    to: str = "",
    status: str = "all",
    facade: str = "all",
):
    if format not in ("pdf", "xlsx"):
        raise HTTPException(status_code=422, detail="Formato inválido (pdf|xlsx)")
    data = await build_report(from_, to, status, facade)
    if format == "pdf":
        content = make_pdf(data, from_, to, status, facade)
        media = "application/pdf"
    else:
        content = make_xlsx(data, from_, to, status, facade)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = f"reporte_{from_}_{to}.{format}"
    return Response(
        content=content,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
