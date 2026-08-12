"""Backend tests for GET /api/photos and display_name dedup in exports (iter 3)."""
import io
import os
import re
import pytest
import requests
from datetime import datetime, timezone

FE_ENV_PATH = "/app/frontend/.env"
BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if os.path.exists(FE_ENV_PATH):
    with open(FE_ENV_PATH) as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
API = f"{BASE_URL}/api"

PNG = bytes.fromhex(
    "89504E470D0A1A0A0000000D49484452000000010000000108060000001F15C4"
    "890000000D49444154789C6360F84F0400000501010A2E7B99790000000049454E44AE426082"
)

UNIQUE = f"TEST_photos_iter3_{datetime.now(timezone.utc).strftime('%H%M%S%f')}"


@pytest.fixture(scope="module")
def s():
    return requests.Session()


def _pick_facade_obj(s, facade="norte"):
    r = s.get(f"{API}/objects", params={"facade": facade, "limit": 20}).json()
    # Prefer objects whose name has the duplicated mark pattern "X X [..]"
    for it in r["items"]:
        parts = it["name"].split(" ")
        if len(parts) >= 2 and parts[0] == parts[1] and it["mark"][0] in ("C", "L"):
            return it["name"]
    for it in r["items"]:
        if it["mark"][0] in ("C", "L"):
            return it["name"]
    return r["items"][0]["name"] if r["items"] else None


# ---------- /api/photos ----------

def test_photos_endpoint_shape(s):
    r = s.get(f"{API}/photos")
    assert r.status_code == 200
    d = r.json()
    assert "total" in d and "items" in d
    assert isinstance(d["items"], list)
    assert isinstance(d["total"], int)


def test_photos_invalid_facade_422(s):
    r = s.get(f"{API}/photos", params={"facade": "bogus"})
    assert r.status_code == 422


def test_photos_creates_and_lists_with_filters(s):
    name = _pick_facade_obj(s, "norte")
    assert name, "no facade object available"
    # Cleanup
    s.put(f"{API}/tags", json={"object_name": name, "status": None, "observation": ""})
    # Upload photo
    up = s.post(f"{API}/upload", files={"file": ("o.png", io.BytesIO(PNG), "image/png")}).json()
    photo_path = up["path"]
    # Save tag with photo + observation
    r = s.put(f"{API}/tags", json={
        "object_name": name,
        "status": "instalado",
        "observation": UNIQUE,
        "photo": photo_path,
    })
    assert r.status_code == 200

    # facade=all
    d = s.get(f"{API}/photos").json()
    matches = [it for it in d["items"] if it["photo"] == photo_path]
    assert matches, "uploaded photo not present in /api/photos"
    it = matches[0]
    for k in ("name", "mark", "facade", "photo", "text", "date", "status"):
        assert k in it
    assert it["name"] == name
    assert it["text"] == UNIQUE
    assert it["status"] == "instalado"
    assert it["facade"] == "norte"

    # facade=norte filter
    dn = s.get(f"{API}/photos", params={"facade": "norte"}).json()
    assert any(it["photo"] == photo_path for it in dn["items"])
    # facade=sur must not include this norte photo
    dsur = s.get(f"{API}/photos", params={"facade": "sur"}).json()
    assert not any(it["photo"] == photo_path for it in dsur["items"])

    # from/to filter: use today's date
    today = datetime.now(timezone.utc).date().isoformat()
    dfrom = s.get(f"{API}/photos", params={"from": today, "to": today}).json()
    assert any(it["photo"] == photo_path for it in dfrom["items"])

    # from in future -> excludes
    future = "2099-01-01"
    dfut = s.get(f"{API}/photos", params={"from": future}).json()
    assert not any(it["photo"] == photo_path for it in dfut["items"])

    # Save the name for xlsx test
    pytest.tag_name = name


# ---------- Dedup in xlsx export ----------

def test_xlsx_uses_display_name(s):
    from openpyxl import load_workbook
    today = datetime.now(timezone.utc).date().isoformat()
    r = s.get(f"{API}/report/export", params={"format": "xlsx", "from": today, "to": today})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Collect all "Pieza" values (col A after header row)
    piezas = []
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] == "Pieza":
            header_row = i
            continue
        if header_row and row and row[0]:
            piezas.append(str(row[0]))
    # No cell should have duplicated mark like "C1 C1 [..]"
    dup_re = re.compile(r"^(\S+)\s+\1\s")
    bad = [p for p in piezas if dup_re.match(p)]
    assert not bad, f"xlsx contains non-deduped names: {bad}"


# ---------- Regression: /api/objects still returns original names ----------

def test_objects_api_keeps_original_names(s):
    d = s.get(f"{API}/objects", params={"limit": 200}).json()
    # There should exist at least one object with duplicated mark pattern
    has_dup = any(
        len(it["name"].split(" ")) >= 2 and it["name"].split(" ")[0] == it["name"].split(" ")[1]
        for it in d["items"]
    )
    assert has_dup, "API should preserve original duplicated names (dedup is UI-only)"


# ---------- Cleanup ----------

def test_cleanup(s):
    name = getattr(pytest, "tag_name", None)
    if name:
        s.put(f"{API}/tags", json={"object_name": name, "status": None, "observation": ""})
